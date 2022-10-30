#판다스 라이브러리로 값을 엑셀로 저장 후 불러오는 코드
from openpyxl import load_workbook

load_wb = load_workbook(r"/workspaces/python_40projects/automation_programs/엑셀의 정보를 불러와 수료증 자동 생성/수료증명단.xlsx")
load_ws = load_wb.active

name_list = []
birthday_list = []
ho_list = []
for i in range(1,load_ws.max_row + 1):
    name_list.append(load_ws.cell(i,1).value)
    birthday_list.append(load_ws.cell(i,2).value)
    ho_list.append(load_ws.cell(i, 3).value)

print(name_list)
print(birthday_list)
print(ho_list)
